from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect as HttpRD, Http404
from django.urls import reverse
from django.shortcuts import render
from django.contrib.auth import logout
from django.views import generic
from django.conf import settings
import datetime as dt

from siteadmin.models import *
from .models import *
from .utils import GAPIDRIVE
from openpyxl import Workbook, load_workbook

DRIVE_DIR = str(settings.BASE_DIR)+str(settings.STATIC_URL)+'gdrive/'
EXCEL_DIR = str(settings.BASE_DIR)+str(settings.STATIC_URL)+'excel/'


def index(req):
	try:
		eu = ExtenUser.objects.get(user=req.user)
	except ExtenUser.DoesNotExist:
		return render(req, 'base/login.html', {'er': "El usuario no ha sido configurado"})
	else:
		site = ""
		if eu.account_type == "1": site = "alumno"
		elif eu.account_type == "2": site = "docente"
		elif eu.account_type == "3": site = "admin"
		else:
			logout(req)
			return HttpRD(reverse('base:login'))

		return HttpRD(reverse('intranet:'+site))


@login_required
def alumno(req):
	sxu = SchoolXUser.objects.get(user=req.user)
	st = Student.objects.get(user=req.user)
	ctx = {'school': sxu.school, 'student': st}
	return render(req, 'intranet/alumno/index.html', ctx)


@login_required
def docente(req):
	sxu = SchoolXUser.objects.get(user=req.user)
	ctx = {'school': sxu.school}
	return render(req, 'intranet/docente/index.html', ctx)


@login_required
def admin(req):
	sxu = SchoolXUser.objects.get(user=req.user)
	ctx = {'school': sxu.school}
	return render(req, 'intranet/admin/index.html', ctx)


""" Views """


def logout_view(req):
	logout(req)
	return HttpRD(reverse('base:login'))


class TeacherGradeList(generic.ListView):
	model = Grade
	template_name = "intranet/docente/grade_list.html"

	def get_queryset(self):
		return TeacherXGradeXSection.objects.filter(teacher__user=self.request.user)


class TeacherGradeDetail(generic.DetailView):
	queryset = Grade.objects.all()
	template_name = "intranet/docente/grade_detail.html"
	context_object_name = 'grade'

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		sxgxs = StudentXGradeXSection.objects.filter(grade=kwargs.get('object'))
		students = [s.student for s in sxgxs]
		ctx.update({'students': students})
		ctx.update({'courses': Course.objects.all()})
		return ctx


class TeacherStudentList(generic.ListView):
	model = TeacherXGradeXSection
	template_name = "intranet/docente/student_list.html"
	context_object_name = 'students'

	def get_queryset(self):
		txgxs = TeacherXGradeXSection.objects.filter(teacher__user=self.request.user)
		grades = [t.grade for t in txgxs]
		sxgxs = StudentXGradeXSection.objects.filter(grade__in=grades)
		return [s.student for s in sxgxs]

	def get_context_data(self, **kwargs):
		from django.db.models import Avg
		ctx = super().get_context_data(**kwargs)
		students = ctx.get('students')
		# Nota
		_students = list()
		for s in students:
			_courses = list()
			courses = Course.objects.all()
			for c in courses:
				promedio = NotaXTrimester.objects.filter(student=s).values('student').annotate(promedio=Avg('nota'))[0].get("promedio")
				_courses.append({
					'name': c.name,
					'prom': promedio})
			_students.append({
				'pk': s.pk,
				'fullname': s.meta.fullname,
				'courses': _courses})
		ctx.update({'students': _students})
		ctx.update({'courses': courses})
		return ctx


class TeacherClasesList(generic.ListView):
	model = SchoolClass
	template_name = "intranet/docente/clase_list.html"

	def get_queryset(self):
		return TeacherXGradeXSection.objects.filter(teacher__user=self.request.user)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		user = self.request.user
		txgxs = TeacherXGradeXSection.objects.filter(teacher__user=user)
		schedules = list()
		for _t in txgxs:
			scss = SchoolClassSchedule.objects.filter(grade=_t.grade, section=_t.section)
			
			_schedule = [
				["08:20 - 09:00"]+[s.course.name for s in scss.filter(time_from="08:20").order_by('dt_day')],
				["09:00 - 09:40"]+[s.course.name for s in scss.filter(time_from="09:00").order_by('dt_day')],
				["09:40 - 10:20"]+[s.course.name for s in scss.filter(time_from="09:40").order_by('dt_day')],
				["10:20 - 11:00"]+[s.course.name for s in scss.filter(time_from="10:20").order_by('dt_day')],
				["11:00 - 11:40"]+[s.course.name for s in scss.filter(time_from="11:00").order_by('dt_day')],
				["11:40 - 12:20"]+[s.course.name for s in scss.filter(time_from="11:40").order_by('dt_day')]]
			schedules.append({
				'grade': _t.grade.pk,
				'section': _t.section.pk,
				'schedule': _schedule})
		ctx.update({'schedules': schedules})
		return ctx


class TeacherClassDetail(generic.DetailView):
	model = SchoolClass
	template_name = "intranet/docente/clase_detail.html"
	context_object_name = 'class'

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		clase = kwargs.get('object')
		files = list()
		if clase.gdrive_folder_id:
			gadrive = GAPIDRIVE()
			files = gadrive.get_files_from_folder(clase.gdrive_folder_id)
		ctx.update({'files': files})
		return ctx


class TeacherCapacitacionList(generic.ListView):
	model = Grade
	template_name = "intranet/docente/capacitacion_list.html"

	def get_queryset(self):
		return TeacherXGradeXSection.objects.filter(teacher__user=self.request.user)


class TeacherGradeCursoDetail(generic.DetailView):
	model = Course
	template_name = "intranet/docente/curso_detail.html"
	context_object_name = 'course'

	def get_object(self, **kwargs):
		return Course.objects.get(pk=self.kwargs.get('pk'))

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		course = ctx.get("course")
		grado_pk = self.kwargs.get("grado")
		scs = SchoolClass.objects.filter(course=course, grade=grado_pk).order_by('dt')
		ctx.update({
			'clases_past': scs.filter(dt__lt=dt.date.today()),
			'clases_today': scs.filter(dt=dt.date.today()),
			'clases_future': scs.filter(dt__gt=dt.date.today()),
			})
		return ctx


class TeacherDetail(generic.DetailView):
	queryset = Teacher.objects.all()
	template_name = "intranet/docente/detail.html"
	context_object_name = 'teacher'


class StudentDetail(generic.DetailView):
	queryset = Student.objects.all()
	template_name = "intranet/alumno/detail.html"
	context_object_name = 'student'

	def get_context_data(self, **kwargs):
		from django.db.models import Avg
		ctx = super().get_context_data(**kwargs)
		student = ctx.get('student')
		# Nota
		nxts = NotaXTrimester.objects.filter(student=student, sc_trimester__sc_year__dt_year__year=2021)
		courses = list()
		for c in Course.objects.all():
			comps = list()
			for cxc in CompetenciaXCurso.objects.filter(course=c):
				nxts = NotaXTrimester.objects.filter(competencia=cxc, student=student)
				notaxtr = [{'order': n.sc_trimester.order, 'nota': n.nota} for n in nxts]
				proms = nxts.values('student', 'competencia').annotate(promedio=Avg('nota'))
				comps.append({
					'name': cxc.competencia,
					'trimester': notaxtr,
					'prom': proms[0].get('promedio')})
			ar_prom = [_c.get("prom") for _c in comps]
			ar_prom = sum(ar_prom)/len(ar_prom)
			courses.append({
				'name': c.name,
				'competencias': comps,
				'prom': ar_prom})
		ctx.update({'courses': courses})
		# Attendance
		atts = Attendance.objects.filter(student=student, sc_trimester__sc_year__dt_year__year=2021)
		ctx.update({'attendances1': atts.filter(sc_trimester__order=1)})
		ctx.update({'attendances2': atts.filter(sc_trimester__order=2)})
		ctx.update({'attendances3': atts.filter(sc_trimester__order=3)})
		return ctx


class StudentCursoDetail(generic.DetailView):
	model = Student
	template_name = "intranet/alumno/curso_list.html"
	context_object_name = 'student'

	def get_object(self, **kwargs):
		return Student.objects.get(user=self.request.user)

	def get_context_data(self, **kwargs):
		ctx = super().get_context_data(**kwargs)
		student = ctx.get("student")
		sxgxs = StudentXGradeXSection.objects.get(student=student)
		courses = Course.objects.all()
		ctx.update({"courses": courses, "grade": sxgxs.grade})
		print(ctx)
		return ctx


""" REPORTE EXCEL """

def ReporteComprobantes(request, student_pk):
	from django.db.models import Avg
	libro = Workbook()
	libro = load_workbook(EXCEL_DIR+"template_reporte_notas_asistencia.xlsx")
	try:
		student = Student.objects.get(pk=student_pk)
	except:
		return Http404("Alumno no encontrado")
	# Notas
	dt_year = 2021
	h1 = libro.get_sheet_by_name("NOTAS")
	h1.cell(row=2, column=2).value = 'REGISTRO DE NOTAS - %s'%dt_year
	h1.cell(row=3, column=2).value = student.meta.fullname
	y = 5
	merge_o = 0
	for c in Course.objects.all():
		merge_o = y+1
		for cxc in CompetenciaXCurso.objects.filter(course=c):
			y += 1
			nxts = NotaXTrimester.objects.filter(competencia=cxc, student=student.pk, sc_trimester__sc_year__dt_year__year=dt_year)
			proms = nxts.values('student', 'competencia').annotate(promedio=Avg('nota'))

			h1.cell(row=y, column=3).value = cxc.competencia
			h1.cell(row=y, column=4).value = round(nxts[0].nota, 1)
			h1.cell(row=y, column=5).value = round(nxts[1].nota, 1)
			h1.cell(row=y, column=6).value = round(nxts[2].nota, 1)
			h1.cell(row=y, column=7).value = round(proms[0].get('promedio'), 1)
		c_prom = NotaXTrimester.objects\
			.filter(competencia__course=c, student=student.pk, sc_trimester__sc_year__dt_year__year=dt_year)\
			.values('student', 'competencia__course')\
			.annotate(promedio=Avg('nota'))[0].get("promedio")

		h1.cell(row=merge_o, column=2).value = c.name
		h1.cell(row=merge_o, column=8).value = round(c_prom, 1)
		h1.merge_cells(start_row=merge_o, start_column=2, end_row=y, end_column=2)
		h1.merge_cells(start_row=merge_o, start_column=8, end_row=y, end_column=8)

	# Attendance
	h2 = libro.get_sheet_by_name("ASISTENCIA")
	h2.cell(row=2, column=2).value = 'REGISTRO DE ASISTENCIA - %s'%dt_year
	h2.cell(row=3, column=2).value = student.meta.fullname

	atts = Attendance.objects.filter(student=student.pk, sc_trimester__sc_year__dt_year__year=2021)
	x = 2
	for at in atts.filter(sc_trimester__order=1):
		x += 1
		h2.cell(row=5, column=x).value = at.dt.strftime('%d-%m-%Y')
		h2.cell(row=6, column=x).value = "ASISTIO" if at.status else "FALTO"
	x = 2
	for at in atts.filter(sc_trimester__order=2):
		x += 1
		h2.cell(row=7, column=x).value = at.dt.strftime('%d-%m-%Y')
		h2.cell(row=8, column=x).value = "ASISTIO" if at.status else "FALTO"

	x = 2
	for at in atts.filter(sc_trimester__order=3):
		x += 1
		h2.cell(row=9, column=x).value = at.dt.strftime('%d-%m-%Y')
		h2.cell(row=10, column=x).value = "ASISTIO" if at.status else "FALTO"


	# Armar respuesta http
	response = HttpResponse(content_type='application/vnd.ms-excel')
	response['Content-Disposition'] = 'attachment; filename=reporte_notas_y_asistencia.xls'
	libro.save(response)
	return response


""" Helpers """


def gdrive_create(req, pk):
	gadrive = GAPIDRIVE()
	folder_id = gadrive.create_folder("class_"+str(pk))
	print("folder_id: %s"%folder_id)
	SchoolClass.objects.filter(pk=pk).update(gdrive_folder_id=folder_id)
	return HttpRD(reverse('intranet:dc_class_detail', args=[pk]))

def gdrive_upload(req, pk):
	gadrive = GAPIDRIVE()
	clase = SchoolClass.objects.get(pk=pk)
	file = req.FILES['file']
	file_path = DRIVE_DIR+file.name
	with open(file_path, 'wb') as f:
		f.write(file.read())
	gadrive_file = gadrive.create_file(file_path, file.name, clase.gdrive_folder_id)
	return HttpRD(reverse('intranet:dc_class_detail', args=[pk]))

def gdrive_delete(req, pk):
	gadrive = GAPIDRIVE()
	file_id = req.POST.get("file_id")
	gadrive.delete_file(file_id)
	return HttpRD(reverse('intranet:dc_class_detail', args=[pk]))

